"""Datatable exporters — CSV / XLSX / PDF.

CSV uses the stdlib `csv` module (no deps).
XLSX uses openpyxl (added as a runtime dep in Phase 11 commit 2).
PDF uses WeasyPrint via apps.invoices.pdf — already a dep for invoices.

PDF export is capped at TABLE_PDF_ROW_CAP rows (per the Phase 11 spec
open-question resolution); CSV and XLSX have no cap.

Each exporter formats values via:
1. col.formatter(value) if set
2. else dotted_attr(row, col.key) → str()

Templates aren't honored — exports are tabular, plain values only. If a
column is purely template-driven (e.g. avatar+name+email composite cell),
add a `formatter=lambda r: r.email` so the export gets a meaningful value.
"""
from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Sequence

from django.db.models import QuerySet
from django.http import HttpResponse
from django.utils import timezone

from apps.core.tables.config import Column

TABLE_PDF_ROW_CAP = 500


def _resolve_value(row, col: Column):
    """Get one cell's exportable value as a string."""
    from apps.core.templatetags.apex import dotted_attr
    raw = dotted_attr(row, col.key)
    if col.formatter is not None:
        try:
            return col.formatter(raw)
        except Exception:
            try:
                return col.formatter(row)
            except Exception:
                pass
    if raw is None or raw == "":
        return ""
    return str(raw)


def _resolve_columns(columns: Iterable[Column]) -> Sequence[Column]:
    return tuple(columns)


def _disposition(filename: str) -> str:
    return f'attachment; filename="{filename}"'


# ── CSV ────────────────────────────────────────────────────────────────


def to_csv(queryset: QuerySet, columns: Iterable[Column], *, filename: str) -> HttpResponse:
    cols = _resolve_columns(columns)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c.label for c in cols])
    for row in queryset.iterator(chunk_size=500):
        writer.writerow([_resolve_value(row, c) for c in cols])
    response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = _disposition(filename)
    return response


# ── XLSX ───────────────────────────────────────────────────────────────


def to_xlsx(queryset: QuerySet, columns: Iterable[Column], *, filename: str) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    cols = _resolve_columns(columns)
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Header row, bold.
    ws.append([c.label for c in cols])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in queryset.iterator(chunk_size=500):
        ws.append([_resolve_value(row, c) for c in cols])

    # Approximate column widths from the longest value in the first ~50 rows.
    for col_idx, _col in enumerate(cols, start=1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row, 50) + 1)),
            default=10,
        )
        ws.column_dimensions[column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = _disposition(filename)
    return response


# ── PDF ────────────────────────────────────────────────────────────────


class PDFRowCapExceeded(Exception):
    """Raised when the queryset for a PDF export exceeds TABLE_PDF_ROW_CAP."""


def to_pdf(queryset: QuerySet, columns: Iterable[Column], *,
           filename: str, title: str) -> HttpResponse:
    """Render the table as a PDF via WeasyPrint.

    Capped at TABLE_PDF_ROW_CAP rows. The cap is checked before WeasyPrint
    is imported so the over-cap response works in environments without
    cairo/pango installed.
    """
    from django.http import HttpResponseBadRequest

    cols = _resolve_columns(columns)
    count = queryset.count()
    if count > TABLE_PDF_ROW_CAP:
        return HttpResponseBadRequest(
            f"PDF export is capped at {TABLE_PDF_ROW_CAP} rows "
            f"(this query has {count}). Filter further or use CSV/XLSX."
        )

    # Lazy import — see apps/invoices/pdf.py for the same pattern.
    from django.template.loader import render_to_string
    from weasyprint import HTML

    rows = list(queryset[:TABLE_PDF_ROW_CAP])
    rendered_rows = [[_resolve_value(r, c) for c in cols] for r in rows]
    html = render_to_string("core/tables/_export.html", {
        "title": title,
        "columns": cols,
        "rows": rendered_rows,
        "now": timezone.now(),
        "row_count": count,
    })
    pdf_bytes = HTML(string=html).write_pdf()
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = _disposition(filename)
    return response
